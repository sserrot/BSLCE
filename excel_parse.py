from openpyxl import load_workbook  # import openpyxl for excel manipulation
import csv_to_xlsx
import time
# RETURNS A LIST OF IDS OF PROGRAM MANAGERS THAT DID NOT ATTEND


def read_attendance(file):
    date = time.strftime("%d-%m-%Y")
    ext = file.split('.')[1]
    if ext == 'csv':
        csv_to_xlsx.main(file)
        file_base = file.split('.')[0]
        filename = file_base + date + '.xlsx'
        file = filename
    wb = load_workbook(file)
    sheet = wb.active
    col = 3  # Default ID column is 3
    attendant_id = []
    pm_id = ''
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).upper() == 'BENTLEY ID':
                col = cell.col_idx

            if cell.col_idx == col and cell.row != 1:  # ignore first row of headings
                pm_id = str(sheet.cell(row=cell.row, column=col).value)
                attendant_id.append(pm_id)
    return attendant_id

def read_master(file):
    date = time.strftime("%d-%m-%Y")
    ext = file.split('.')[1]
    if ext == 'csv':
        csv_to_xlsx.main(file)
        file_base = file.split('.')[0]
        filename = file_base + date + '.xlsx'
        file = filename
    wb = load_workbook(file)
    sheet = wb.active
    master_list = []
    pm_id = ''
    first_name = ''
    last_name = ''
    email = ''

    col_id = 1  # Default ID column is 1 etc...
    col_fn = 2
    col_ln = 3
    col_e = 4

    for row in sheet.iter_rows():
        information = {}
        for cell in row:
            if str(cell.value).upper() == 'BENTLEY ID':
                col_id = cell.col_idx
            if cell.col_idx == col_id and cell.row != 1:  # ignore first row of headings
                pm_id = str(sheet.cell(row=cell.row, column=col_id).value)

            if str(cell.value).upper() == 'FIRSTNAME':
                col_fn = cell.col_idx
            if cell.col_idx == col_fn and cell.row != 1:  # ignore first row of headings
                first_name = str(sheet.cell(row=cell.row, column=col_fn).value)

            if str(cell.value).upper() == 'LASTNAME':
                col_ln = cell.col_idx
            if cell.col_idx == col_ln and cell.row != 1:  # ignore first row of headings
                last_name = str(sheet.cell(row=cell.row, column=col_ln).value)

            if str(cell.value).upper() == 'EMAIL':
                col_e = cell.col_idx
            if cell.col_idx == col_e and cell.row != 1:  # ignore first row of headings
                email = str(sheet.cell(row=cell.row, column=col_e).value)

            information['first_name'] = first_name
            information['pm_id'] = pm_id
            information['last_name'] = last_name
            information['email'] = email

        master_list.append(information)

    del master_list[0]  # delete the first row of empties
    return master_list
