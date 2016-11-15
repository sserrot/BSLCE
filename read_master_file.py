from openpyxl import load_workbook  # import openpyxl for excel manipulation

# RETURNS A LIST OF DICTIONARIES CONTAINING ALL PROGRAM MANAGERS INFORMATIONS

wb = load_workbook('master.xlsx')
sheet = wb.get_sheet_by_name(name='Sheet1')


def master():
    master_list = []
    pm_id = ''
    first_name = ''
    last_name = ''
    email = ''

    col_id = 1  # Default ID column is 1 etc...
    col_fn = 2
    col_ln = 3
    col_e = 4

    for row in sheet.iter_rows():
        information = {}
        if row[0].value is not None:
            for cell in row:
                if str(cell.value).upper() == 'ID':
                    col_id = cell.col_idx
                if cell.col_idx == col_id and cell.row != 1:  # ignore first row of headings
                    pm_id = str(sheet.cell(row=cell.row, column=col_id).value)

                if str(cell.value).upper() == 'FIRST NAME':
                    col_fn = cell.col_idx
                if cell.col_idx == col_fn and cell.row != 1:  # ignore first row of headings
                    first_name = str(sheet.cell(row=cell.row, column=col_fn).value)

                if str(cell.value).upper() == 'LAST NAME':
                    col_ln = cell.col_idx
                if cell.col_idx == col_ln and cell.row != 1:  # ignore first row of headings
                    last_name = str(sheet.cell(row=cell.row, column=col_ln).value)

                if str(cell.value).upper() == 'EMAIL':
                    col_e = cell.col_idx
                if cell.col_idx == col_e and cell.row != 1:  # ignore first row of headings
                    email = str(sheet.cell(row=cell.row, column=col_e).value)

                information['first_name'] = first_name
                information['pm_id'] = pm_id
                information['last_name'] = last_name
                information['email'] = email

            master_list.append(information)
        else:
            break
    del master_list[0]  # delete the first row of empties
    return master_list