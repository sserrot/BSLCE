from openpyxl import load_workbook # import openpyxl for excel manipulation
from openpyxl.utils import column_index_from_string
# TO DO - abstract column 4 part of code and (D) part

nonAttendantStudents = {}
wb = load_workbook('attendance.xlsx')
sheet = wb.get_sheet_by_name(name='Sheet1')
col = 2  # Default ID column is 3

for row in sheet.iter_rows():
    if row[0].value is not None:
        for cell in row:
            if str(cell.value).upper() == 'ID':
                col = cell.col_idx
            print(cell.value)
            ID = str(sheet.cell(row=cell.row, column=col).value)
            if attendance == 'MISS':  # check if person attended
                full_name = sheet.cell(row=cell.row, column=1).value   # store full name
                last = full_name.rsplit(',')[0].strip(',')  # format - broken if MISS is empty
                formatted_last = last.replace(" ", "")  # get of of spaces
                formatted_last = formatted_last.replace("\'", "")  # get rid of '
                if len(full_name.rsplit(' ')) == 1:
                        continue
                else:
                    first = full_name.rsplit(',')[1]  # format, causes an error if not in LAST, FIRST format
                    first = first.strip()
                    # email format is 7last_2first@bentley.edu
                    email = (formatted_last[:7] + '_' + first[:4] + '@bentley.edu')
                    name = first + ' ' + last
                    nonAttendantStudents[name] = email  # add to a dictionary
                    continue
            else:
                continue
    # except InsufficientCoordinatesException:
    else:
        break

non_attend_names = ' '
for name in nonAttendantStudents.items():
    non_attend_names = non_attend_names + ' ' + name[0]
    print(non_attend_names)

