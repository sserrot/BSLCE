import csv
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def main(filepath):
    f = open(filepath)
    date = time.strftime("%d-%m-%Y")

    csv.register_dialect('comma', delimiter=',')
    reader = csv.reader(f, dialect='comma')
    dest_file_base = filepath.split('.')[0]
    dest_filename = dest_file_base + date + '.xlsx'
    
    wb = Workbook()

    ws = wb.worksheets[0]

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

    wb.save(filename = dest_filename)