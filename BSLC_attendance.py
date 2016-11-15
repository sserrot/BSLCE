from openpyxl import load_workbook # import openpyxl for excel manipulation
from openpyxl.utils.exceptions import (
    InsufficientCoordinatesException
)
import smtplib  # import smtplib to send email
# TODO - abstract column 2 part of code

nonAttendantStudents = {}


def excel_parse(file):
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name(name='Sheet1')
    for row in sheet.iter_rows():
        # try:  # stop when no more cells
        if row[0].value is not None:
            for cell in row:
                attendance = str(sheet.cell(row=cell.row, column=2).value).upper()
                if attendance == 'MISS':  # check if person attended
                    full_name = sheet.cell(row=cell.row, column=1).value  # store full name
                    last = full_name.rsplit(',')[0].strip(',')  # format - broken if MISS is empty
                    formatted_last = last.replace(" ", "")  # get of of spaces
                    formatted_last = formatted_last.replace("\'", "")  # get rid of '
                    if len(full_name.rsplit(' ')) == 1:
                        continue
                    else:
                        first = full_name.rsplit(',')[1]  # format, causes an error if not in LAST, FIRST format
                        first = first.strip()
                        # email format is 7last_2first@bentley.edu
                        email = (formatted_last[:7] + '_' + first[:4] + '@bentley.edu')
                        name = first + ' ' + last
                        nonAttendantStudents[name] = email  # add to a dictionary
                        continue
                else:
                    continue
        # except InsufficientCoordinatesException:
        else:
            break

    non_attend_names = ' '
    for name in nonAttendantStudents.items():
        non_attend_names = non_attend_names + ' ' + name[0]
        print(non_attend_names)
# create a text/plain message


def send_emails(sender_address, password, subject, text_body):

    sender_address = sender_address
    password = password
    subject = subject
    text_body = text_body

    smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
    smtp_obj.ehlo()
    smtp_obj.starttls()
    smtp_obj.login(sender_address, password)

# calling str on text_body creates problems with unicode ... TODO - fix this
    for name, email in nonAttendantStudents.items():
        body = ("Subject:" + str(subject) + "\n" + "Dear {}, \n\n" + str(text_body)).format(name)
        print('Sending email to %s...' % email)
        send_mail_status = smtp_obj.sendmail(sender_address, email, body)

        if send_mail_status != {}:
            print('There was a problem sending email to %s: %s' % (email, send_mail_status))
    all_body = "Subject:" + str(subject) + "\n" + "Dear BSLCE, \n\n"
    for name in nonAttendantStudents.items():
        all_body = all_body + ' ' + name[0] + ";"

    send_all_status = smtp_obj.sendmail(sender_address, sender_address, all_body)

    if send_all_status != {}:
        print('There was a problem sending email to %s: %s' % (sender_address,send_all_status))
    smtp_obj.quit()


def main(file, sender_address, password, subject, text_body):
    excel_parse(file=file)
    send_emails(sender_address=sender_address, password=password, subject=subject, text_body=text_body)

if __name__ == '__main__':
    main(file='foo.xlsx', sender_address='no@gmail.com', password='pass', subject='subj', text_body='text_body')
