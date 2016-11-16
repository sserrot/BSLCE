from openpyxl import load_workbook  # import openpyxl for excel manipulation

# RETURNS A LIST OF IDS OF PROGRAM MANAGERS THAT DID NOT ATTEND


def attended(file):
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name(name='Sheet1')
    col = 3  # Default ID column is 3
    attendant_id = []
    pm_id = ''
    for row in sheet.iter_rows():
        if row[0].value is not None:
            for cell in row:
                if str(cell.value).upper() == 'ID':
                    col = cell.col_idx

                if cell.col_idx == col and cell.row != 1:  # ignore first row of headings
                    pm_id = str(sheet.cell(row=cell.row, column=col).value)
                    attendant_id.append(pm_id)
        else:
            break
    return attendant_id

